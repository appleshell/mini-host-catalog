#!/usr/bin/env python3
"""
Excel 产品数据导入脚本
将 产品分类选型表.xlsx 解析后写入 SQLite 数据库，同时提取图片到 api/public/images/
"""

import sqlite3
import re
import os
import zipfile
import xml.etree.ElementTree as ET
import openpyxl

_BASE      = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(_BASE, '../产品分类选型表.xlsx')
DB_PATH    = os.path.join(_BASE, '../api/data/catalog.db')
IMG_DIR    = os.path.join(_BASE, '../api/public/images')

NS = {
    'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
    'a':   'http://schemas.openxmlformats.org/drawingml/2006/main',
    'r':   'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
    'rel': 'http://schemas.openxmlformats.org/package/2006/relationships',
}


# ── 辅助函数 ──────────────────────────────────────────────

def parse_network_count(text: str) -> int:
    if not text:
        return 1
    matches = re.findall(r'(\d+)\*.*?网', text)
    return sum(int(m) for m in matches) if matches else 1


def parse_serial_count(text: str) -> int:
    if not text:
        return 0
    matches = re.findall(r'(\d+)\*.*?(?:RS232|RS485|串口|COM)', text, re.IGNORECASE)
    return sum(int(m) for m in matches) if matches else 0


def has_wifi(text: str) -> bool:
    return bool(text) and ('WIFI' in text.upper() or 'WI-FI' in text.upper())


def has_fanless(text: str) -> bool:
    return bool(text) and '无风扇' in text


def clean(val) -> str:
    if val is None:
        return ''
    return str(val).strip().replace('\xa0', ' ').replace('\n', ' ').replace('\r', '')


# ── 图片提取 ──────────────────────────────────────────────

def extract_images(xlsx_path: str) -> dict[str, dict[int, list[str]]]:
    """
    返回 { sheet_index(1-based): { row(0-based): [image_filename, ...] } }
    row 是 drawing.xml 里的 <xdr:from><xdr:row> 值（0-based）
    """
    result: dict[int, dict[int, list[str]]] = {}

    with zipfile.ZipFile(xlsx_path, 'r') as z:
        all_files = z.namelist()

        for sheet_idx in range(1, 10):
            rels_path    = f'xl/worksheets/_rels/sheet{sheet_idx}.xml.rels'
            if rels_path not in all_files:
                continue

            # 找到该 sheet 对应的 drawing 编号
            rels_root = ET.fromstring(z.read(rels_path).decode())
            drawing_target = None
            for rel in rels_root.findall('rel:Relationship', NS):
                if 'drawing' in rel.get('Type', ''):
                    # Target 形如 "../drawings/drawing1.xml"
                    drawing_target = rel.get('Target', '').split('/')[-1]
                    break
            if not drawing_target:
                continue

            drawing_num  = re.search(r'\d+', drawing_target).group()
            drawing_path = f'xl/drawings/drawing{drawing_num}.xml'
            drwrel_path  = f'xl/drawings/_rels/drawing{drawing_num}.xml.rels'

            if drawing_path not in all_files:
                continue

            # rId → 图片文件名
            rid_to_file: dict[str, str] = {}
            if drwrel_path in all_files:
                drwrel_root = ET.fromstring(z.read(drwrel_path).decode())
                for rel in drwrel_root.findall('rel:Relationship', NS):
                    rid  = rel.get('Id')
                    tgt  = rel.get('Target', '')          # ../media/imageN.png
                    fname = tgt.split('/')[-1]            # imageN.png
                    rid_to_file[rid] = fname

            # 解析 drawing.xml：锚点 row → rId
            drw_root = ET.fromstring(z.read(drawing_path).decode())
            row_to_images: dict[int, list[str]] = {}

            for anchor in drw_root.findall('xdr:twoCellAnchor', NS) + drw_root.findall('xdr:oneCellAnchor', NS):
                from_el = anchor.find('xdr:from', NS)
                if from_el is None:
                    continue
                row_el = from_el.find('xdr:row', NS)
                if row_el is None:
                    continue
                row = int(row_el.text)   # 0-based

                blip = anchor.find('.//a:blip', NS)
                if blip is None:
                    continue
                rid = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                if rid and rid in rid_to_file:
                    row_to_images.setdefault(row, []).append(rid_to_file[rid])

            result[sheet_idx] = row_to_images

    return result


def copy_images(img_filenames: list[str], model: str, z: zipfile.ZipFile) -> list[str]:
    """把图片从 zip 直接读出写到 IMG_DIR/{model}/，返回 URL 路径列表"""
    dest_dir = os.path.join(IMG_DIR, model)
    os.makedirs(dest_dir, exist_ok=True)

    all_files = z.namelist()
    urls = []
    for i, fname in enumerate(img_filenames, 1):
        src = f'xl/media/{fname}'
        ext = os.path.splitext(fname)[1]
        dest = os.path.join(dest_dir, f'{i}{ext}')
        if src in all_files:
            with open(dest, 'wb') as f:
                f.write(z.read(src))
            urls.append(f'/images/{model}/{i}{ext}')
    return urls


# ── 数据库迁移 ────────────────────────────────────────────

def run_migrations(conn: sqlite3.Connection):
    migration_dir = os.path.join(os.path.dirname(__file__), '../api/drizzle')
    sql_files = sorted(f for f in os.listdir(migration_dir) if f.endswith('.sql'))
    for sql_file in sql_files:
        with open(os.path.join(migration_dir, sql_file), 'r') as f:
            sql = f.read()
        sql = sql.replace('CREATE TABLE ', 'CREATE TABLE IF NOT EXISTS ')
        conn.executescript(sql)
    print(f'✓ 执行了 {len(sql_files)} 个迁移文件')


# ── 主流程 ────────────────────────────────────────────────

def import_products():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    os.makedirs(IMG_DIR, exist_ok=True)

    conn = sqlite3.connect(DB_PATH)
    conn.execute('PRAGMA journal_mode=WAL')
    run_migrations(conn)
    conn.execute('DELETE FROM products')
    conn.execute("DELETE FROM sqlite_sequence WHERE name='products'")

    # 预先解析全部图片映射
    print('📦 解析图片映射...')
    sheet_img_map = extract_images(EXCEL_PATH)

    wb   = openpyxl.load_workbook(EXCEL_PATH)
    total = 0

    with zipfile.ZipFile(EXCEL_PATH, 'r') as z:
        for sheet_idx, sheet_name in enumerate(wb.sheetnames, 1):
            ws          = wb[sheet_name]
            row_img_map = sheet_img_map.get(sheet_idx, {})
            count       = 0

            # 收集产品行（从第2行开始，index 从1），记录行号
            data_rows = []
            for excel_row_num, row in enumerate(
                ws.iter_rows(min_row=2, values_only=True), start=2
            ):
                if any(row):
                    data_rows.append((excel_row_num, row))

            for excel_row_num, row in data_rows:
                (category, model, _img_col, _size, _ports,
                 cpu, memory, storage, gpu, network,
                 audio, display, other_ports, dimensions,
                 os_text, power, cooling, mounting, notes, *_) = (list(row) + [None] * 20)[:20]

                model = clean(model)
                # 过滤空值和纯符号的无效型号（如 '/'）
                if not model or not re.search(r'[A-Za-z0-9\u4e00-\u9fa5]', model):
                    continue
                # 型号中的斜杠替换为下划线，避免路径问题
                model_safe = model.replace('/', '_').replace('\\', '_')

                # drawing.xml 里 row 是 0-based，Excel 行号从1开始
                # 第2行 Excel → drawing row index = excel_row_num - 1
                drawing_row = excel_row_num - 1
                img_files   = row_img_map.get(drawing_row, [])

                # 提取图片（用 model_safe 作为目录名）
                urls = copy_images(img_files, model_safe, z) if img_files else []
                image_url = urls[0] if urls else ''

                conn.execute('''
                    INSERT INTO products (
                        category, model, image_url, cpu, memory, storage,
                        gpu, network, audio, display, other_ports, dimensions,
                        os, power, cooling, mounting, notes,
                        network_count, serial_count, has_wifi, has_fanless
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    clean(category) or sheet_name,
                    model,
                    image_url,
                    clean(cpu), clean(memory), clean(storage),
                    clean(gpu), clean(network), clean(audio),
                    clean(display), clean(other_ports), clean(dimensions),
                    clean(os_text), clean(power), clean(cooling),
                    clean(mounting), clean(notes),
                    parse_network_count(clean(network)),
                    parse_serial_count(clean(other_ports)),
                    1 if has_wifi(clean(network)) else 0,
                    1 if has_fanless(clean(cooling)) else 0,
                ))
                count += 1

            print(f'  [{sheet_name}] 导入 {count} 条，图片行数 {len(row_img_map)}')
            total += count

    conn.commit()
    conn.close()
    print(f'\n✅ 共导入 {total} 条产品数据 → {DB_PATH}')
    print(f'🖼  图片保存至 → {IMG_DIR}')


if __name__ == '__main__':
    import_products()
